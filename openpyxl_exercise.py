import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb=xl.load_workbook('filename')
    sheet=wb['Sheet1']
    m=0
    for i in range(2,sheet.max_row+1):
        if sheet.cell(i,3).value == None:
            m+=1
            break
        else:
            sheet.cell(i,4).value = sheet.cell(i,3).value * 0.9
            m+=1

    values=Reference(sheet,min_row=2,max_row=m,min_col=3,max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save('filename')